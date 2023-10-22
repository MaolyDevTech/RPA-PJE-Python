from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from time import sleep
import openpyxl

# Número da OAB e estado a serem consultados
numero_oab = 133864
estado = 'SP'

# Nome do arquivo Excel
excel_file = 'dados.xlsx'

# Configuração do WebDriver do Chrome
options = webdriver.ChromeOptions()
options.add_argument('--start-maximized')  # Maximizar a janela do Chrome
driver = webdriver.Chrome(options=options)

# Função para extrair informações e salvar no Excel
def extrair_e_salvar_dados(driver, workbook):
    janelas = driver.window_handles
    driver.switch_to.window(janelas[-1])
    sleep(2)

    # Extrair o número do processo e data da distribuição
    numero_processo = driver.find_element(By.XPATH, "//div[@class='col-sm-12 ']").text
    data_distribuicao = driver.find_elements(By.XPATH, "//div[@class='value col-sm-12 ']")[1].text

    # Extrair e guardar todas as últimas movimentações
    movimentacoes = driver.find_elements(By.XPATH, "//div[@id='j_id132:processoEventoPanel_body']//tr[contains(@class,'rich-table-row')]//td//div//div//span")
    lista_movimentacoes = [movimentacao.text for movimentacao in movimentacoes]

    try:
        # Tentar acessar uma página existente no Excel
        pagina_processo = workbook[numero_processo]
    except KeyError:
        # Criar uma nova página se não existir
        pagina_processo = workbook.create_sheet(numero_processo)
        # Definir o cabeçalho
        pagina_processo['A1'] = "Número do Processo"
        pagina_processo['B1'] = "Data de Distribuição"
        pagina_processo['C1'] = "Movimentações"

    # Adicionar os dados ao Excel
    pagina_processo['A2'] = numero_processo
    pagina_processo['B2'] = data_distribuicao
    for index, linha in enumerate(pagina_processo.iter_rows(min_row=3, max_row=len(lista_movimentacoes) + 2, min_col=3, max_col=3)):
        for celula in linha:
            celula.value = lista_movimentacoes[index]

    workbook.save(excel_file)
    sleep(2)

# Acesse o site
driver.get('https://pje-consulta-publica.tjmg.jus.br/')
sleep(30)

# Digite o número da OAB
campo_oab = driver.find_element(By.XPATH, "//input[@id='fPP:Decoration:numeroOAB']")
campo_oab.send_keys(numero_oab)

# Selecione o estado
dropdown_estados = driver.find_element(By.XPATH, "//select[@id='fPP:Decoration:estadoComboOAB']")
opcoes_estados = Select(dropdown_estados)
opcoes_estados.select_by_visible_text(estado)

# Clique em "Pesquisar"
botao_pesquisar = driver.find_element(By.XPATH, "//input[@id='fPP:searchProcessos']")
botao_pesquisar.click()
sleep(10)

# Entre em cada um dos processos
processos = driver.find_elements(By.XPATH, "//b[@class='btn-block']")
for processo in processos:
    try:
        processo.click()
        sleep(10)
        extrair_e_salvar_dados(driver, openpyxl.load_workbook(excel_file))
    except Exception as error:
        print(f"Erro ao processar um processo: {error}")
        continue

# Encerre o navegador
driver.quit()
